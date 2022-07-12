#!/usr/bin/python3

from openpyxl import load_workbook
import yaml
import os

# 1.打开 Excel 表格并获取表格名称
workbook = load_workbook(filename="/home/li/tool/ist_error.xlsx")
print(workbook.sheetnames)
# 2.通过 sheet 名称获取表格
sheet = workbook["Sheet1"]
print(sheet)

elist = []
# # 获取一系列格子
# # 获取 A1:C2 区域的值
cell = sheet["F2:G451"]
print(cell)
for i in cell:
    k = 0
    edict = {}
    for j in i:
        if k == 0:
            print("ec:",j.value)
            ec = '0x'+j.value
            edict['ec'] = int(ec, 16)
        else:
            print("iec:",j.value)
            iec = '0x'+j.value
            edict['iec'] = int(iec, 16)
            elist.append(edict)
        k = k +1
rdict = {'maps':elist}
yamlpath =  "/home/li/caps.yaml"

# 写入到yaml文件
with open(yamlpath, "w", encoding="utf-8") as f:
    yaml.dump(rdict, f) 

